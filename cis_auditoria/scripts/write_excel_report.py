#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import argparse
import pathlib
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# Colores
HEADER_FILL = "1F4E78"      # Azul oscuro
HEADER_FONT = "FFFFFF"      # Blanco
GREEN_FILL = "C6EFCE"       # Verde claro
RED_FILL = "FCE4D6"         # Rojo/Naranja claro
GRAY_FILL = "E7E6E6"        # Gris claro
YELLOW_FILL = "FFF2CC"      # Amarillo claro


def safe_text(value: Any, max_len: int = 32760) -> str:
    """
    Convierte cualquier valor a texto seguro para Excel.
    Limita longitud para evitar problemas con celdas muy grandes.
    """
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > max_len:
        return text[:max_len] + " ...[TRUNCADO]"
    return text


def autosize(ws) -> None:
    """
    Ajusta automáticamente el ancho de columnas.
    """
    widths: Dict[int, int] = {}

    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            length = len(str(val)) if val is not None else 0
            widths[idx] = max(widths.get(idx, 10), min(length + 2, 80))

    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_header(ws, row_num: int = 1) -> None:
    """
    Da formato al encabezado.
    """
    for cell in ws[row_num]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_area(ws) -> None:
    """
    Aplica wrap text y alineación vertical a toda la hoja.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def normalize_result(item: Dict[str, Any], host: str, ip: str) -> Dict[str, Any]:
    """
    Normaliza cada resultado del JSON para que el Excel sea consistente.
    """
    control_id = (
        item.get("id")
        or item.get("control_id")
        or item.get("CONTROL")
        or ""
    )

    name = (
        item.get("name")
        or item.get("control_name")
        or item.get("DESCRIPCIÓN")
        or item.get("DESCRIPCION")
        or ""
    )

    section = item.get("section", "")
    category = item.get("category", "")
    rationale = item.get("rationale", "")
    exclusions = safe_text(item.get("exclusions", ""))

    command = (
        item.get("command")
        or item.get("validation_command")
        or item.get("COMANDO_DE_VALIDACION")
        or item.get("COMANDO DE VALIDACIÓN")
        or ""
    )

    rc = item.get("rc", "")

    stdout = safe_text(item.get("stdout", ""))
    stderr = safe_text(item.get("stderr", ""))

    status = item.get("status")
    passed = item.get("passed", None)

    if not status:
        # Fallback si no viene "status"
        if isinstance(passed, bool):
            status = "Cumple" if passed else "No cumple"
        else:
            # Intentar inferir por stdout
            if stdout.strip().lower() == "not_applicable":
                status = "No aplicable"
            elif rc == 0:
                status = "Cumple"
            else:
                status = "No cumple"

    evidence = item.get("evidence")
    if not evidence:
        if stdout:
            evidence = stdout
        elif stderr:
            evidence = stderr
        else:
            evidence = f"Sin salida; rc={rc}"

    evidence = safe_text(evidence)

    return {
        "host": host,
        "ip": ip,
        "section": section,
        "control_id": control_id,
        "name": name,
        "category": category,
        "command": command,
        "status": status,
        "evidence": evidence,
        "rc": rc,
        "stdout": stdout,
        "stderr": stderr,
        "rationale": rationale,
        "exclusions": exclusions,
    }


def color_status_cell(cell, status: str) -> None:
    """
    Colorea la celda de resultado según el estado.
    """
    normalized = (status or "").strip().lower()

    if normalized == "cumple":
        cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
    elif normalized == "no cumple":
        cell.fill = PatternFill("solid", fgColor=RED_FILL)
    elif normalized == "no aplicable":
        cell.fill = PatternFill("solid", fgColor=GRAY_FILL)
    else:
        cell.fill = PatternFill("solid", fgColor=YELLOW_FILL)


def build_summary_sheet(wb: Workbook, normalized_results: List[Dict[str, Any]], host: str, ip: str) -> None:
    """
    Crea una hoja resumen.
    """
    ws = wb.create_sheet(title="Resumen", index=0)

    total = len(normalized_results)
    cumple = sum(1 for r in normalized_results if r["status"] == "Cumple")
    no_cumple = sum(1 for r in normalized_results if r["status"] == "No cumple")
    no_aplica = sum(1 for r in normalized_results if r["status"] == "No aplicable")

    rows = [
        ["HOST", host],
        ["IP", ip],
        ["TOTAL CONTROLES", total],
        ["CUMPLEN", cumple],
        ["NO CUMPLEN", no_cumple],
        ["NO APLICAN", no_aplica],
    ]

    for row in rows:
        ws.append(row)

    for cell in ws["A"]:
        cell.font = Font(bold=True)

    style_data_area(ws)
    autosize(ws)


def build_detail_sheet(wb: Workbook, normalized_results: List[Dict[str, Any]]) -> None:
    """
    Crea la hoja detallada de auditoría.
    """
    ws = wb.active
    ws.title = "Auditoria"

    headers = [
        "HOST",
        "IP",
        "SECCION",
        "CONTROL",
        "DESCRIPCION",
        "CATEGORIA",
        "COMANDO DE VALIDACION",
        "RESULTADO",
        "EVIDENCIA / MOTIVO",
        "EXCLUSIONES",
        "RC",
        "STDOUT",
        "STDERR",
        "OBSERVACION",
    ]
    ws.append(headers)
    style_header(ws, 1)

    for result in normalized_results:
        row = [
            result["host"],
            result["ip"],
            result["section"],
            result["control_id"],
            result["name"],
            result["category"],
            result["command"],
            result["status"],
            result["evidence"],
            result["exclusions"],
            result["rc"],
            result["stdout"],
            result["stderr"],
            result["rationale"],
        ]
        ws.append(row)

        # Colorear la columna RESULTADO
        status_cell = ws.cell(row=ws.max_row, column=8)
        color_status_cell(status_cell, result["status"])

    # Congelar encabezado
    ws.freeze_panes = "A2"

    # Autofiltro
    ws.auto_filter.ref = ws.dimensions

    style_data_area(ws)
    autosize(ws)


def build_report(json_path: str, output_path: str) -> None:
    """
    Lee el JSON y genera el Excel.
    """
    data = json.loads(pathlib.Path(json_path).read_text(encoding="utf-8"))

    host = data.get("host", "")
    ip = data.get("ip", "")
    results = data.get("results", [])

    normalized_results = [normalize_result(item, host, ip) for item in results]

    wb = Workbook()
    build_summary_sheet(wb, normalized_results, host, ip)
    build_detail_sheet(wb, normalized_results)

    output_file = pathlib.Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def main() -> None:
    parser = argparse.ArgumentParser(description="Crear XLSX de auditoría CIS por host")
    parser.add_argument("--input", required=True, help="Ruta del JSON con resultados")
    parser.add_argument("--output", required=True, help="Ruta del archivo XLSX de salida")
    args = parser.parse_args()

    build_report(args.input, args.output)


if __name__ == "__main__":
    main()